"""
Reports Router — CompressorAI v5
Professional formal PDF and Excel reports.

FIX: Import get_current_user from deps (not routers.auth)
FIX: /my endpoint uses analysis_results (not analyses)
"""
import io
from datetime import datetime
from typing import Optional, Dict, Any

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from config import get_supabase_client
# FIX: was "from routers.auth import get_current_user" — should use deps
from deps import get_current_user

router = APIRouter()


class ReportRequest(BaseModel):
    compressor_id:    str
    compressor_name:  str
    analysis_results: Dict[str, Any]
    user_params:      Optional[Dict]  = {}
    company_name:     Optional[str]   = "Industrial Facility"
    include_graphs:   Optional[bool]  = True


# ── PDF Report ────────────────────────────────────────────────
@router.post("/pdf")
async def generate_pdf_report(request: ReportRequest, current_user=Depends(get_current_user)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, HRFlowable, KeepTogether)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import Frame, PageTemplate, BaseDocTemplate

        W, H = A4
        buf  = io.BytesIO()

        # ── Colours ──
        C_NAVY     = colors.HexColor("#0a1628")
        C_BLUE     = colors.HexColor("#1a3a6b")
        C_LBLUE    = colors.HexColor("#e8f0fe")
        C_WHITE    = colors.white
        C_CYAN     = colors.HexColor("#0066cc")
        C_GREEN    = colors.HexColor("#15803d")
        C_ORANGE   = colors.HexColor("#c2410c")
        C_GRAY     = colors.HexColor("#64748b")
        C_LGRAY    = colors.HexColor("#f1f5f9")
        C_BORDER   = colors.HexColor("#cbd5e1")
        C_TEXT     = colors.HexColor("#1e293b")
        C_SUBTEXT  = colors.HexColor("#475569")

        # ── Styles ──
        title_s  = ParagraphStyle("title",  fontSize=22, textColor=C_WHITE,   fontName="Helvetica-Bold",  spaceAfter=2,  leading=26)
        sub_s    = ParagraphStyle("sub",    fontSize=10, textColor=colors.HexColor("#94a3b8"), fontName="Helvetica", spaceAfter=0)
        h2_s     = ParagraphStyle("h2",     fontSize=11, textColor=C_WHITE,   fontName="Helvetica-Bold",  spaceAfter=0,  leading=16)
        body_s   = ParagraphStyle("body",   fontSize=9,  textColor=C_TEXT,    fontName="Helvetica",       spaceAfter=4,  leading=14)
        small_s  = ParagraphStyle("small",  fontSize=8,  textColor=C_SUBTEXT, fontName="Helvetica",       spaceAfter=2)
        footer_s = ParagraphStyle("footer", fontSize=7.5,textColor=C_GRAY,    fontName="Helvetica",       alignment=TA_CENTER)
        kpi_val  = ParagraphStyle("kpiv",   fontSize=18, textColor=C_CYAN,    fontName="Helvetica-Bold",  alignment=TA_CENTER, leading=22)
        kpi_lbl  = ParagraphStyle("kpil",   fontSize=7.5,textColor=C_SUBTEXT, fontName="Helvetica",       alignment=TA_CENTER, leading=10)
        kpi_unit = ParagraphStyle("kpiu",   fontSize=7,  textColor=C_GRAY,    fontName="Helvetica",       alignment=TA_CENTER)
        saving_v = ParagraphStyle("savv",   fontSize=22, textColor=C_GREEN,   fontName="Helvetica-Bold",  alignment=TA_CENTER, leading=26)

        r    = request.analysis_results
        up   = request.user_params or {}
        ts   = datetime.now().strftime("%d %B %Y  %H:%M")
        ts_f = datetime.now().strftime("%d %b %Y %H:%M")

        saving    = r.get("power_saving_percent", 0) or 0
        best_elec = r.get("best_electrical_power", 0) or 0
        best_mech = r.get("best_mechanical_power", 0) or 0
        base_elec = r.get("baseline_electrical_power", 0) or 0
        best_spc  = r.get("best_spc", 0) or 0
        scores    = r.get("scores", {})
        opt_p     = r.get("optimal_parameters", {})
        fi        = r.get("feature_importance", {})
        cs        = r.get("cluster_stats", {})

        story = []
        LM = 1.8*cm

        # HEADER BANNER
        logo_data = [[
            Paragraph("⚡", ParagraphStyle("logo", fontSize=28, textColor=colors.HexColor("#facc15"),
                      alignment=TA_CENTER, fontName="Helvetica-Bold")),
            Table([
                [Paragraph("CompressorAI", ParagraphStyle("brand", fontSize=20,
                    textColor=C_WHITE, fontName="Helvetica-Bold", spaceAfter=2))],
                [Paragraph("Industrial Air Compressor Optimizer",
                    ParagraphStyle("brand2", fontSize=9, textColor=colors.HexColor("#94a3b8"),
                    fontName="Helvetica", spaceAfter=0))],
                [Paragraph("v5.0  ·  DBSCAN + GBR + Genetic Algorithm",
                    ParagraphStyle("brand3", fontSize=7.5, textColor=colors.HexColor("#64748b"),
                    fontName="Helvetica"))],
            ], colWidths=[12*cm]),
            Table([
                [Paragraph("ANALYSIS REPORT", ParagraphStyle("rtype", fontSize=9,
                    textColor=colors.HexColor("#facc15"), fontName="Helvetica-Bold",
                    alignment=TA_RIGHT, spaceAfter=4))],
                [Paragraph(ts, ParagraphStyle("rdate", fontSize=8,
                    textColor=colors.HexColor("#94a3b8"), fontName="Helvetica",
                    alignment=TA_RIGHT))],
            ], colWidths=[5*cm]),
        ]]
        logo_t = Table(logo_data, colWidths=[1.6*cm, 12*cm, 5*cm])
        logo_t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), C_NAVY),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0), (0,-1),  14),
            ("RIGHTPADDING", (-1,0),(-1,-1), 14),
            ("TOPPADDING",   (0,0), (-1,-1), 14),
            ("BOTTOMPADDING",(0,0), (-1,-1), 14),
            ("LINEBELOW",    (0,0), (-1,-1), 3, colors.HexColor("#facc15")),
        ]))
        story.append(logo_t)

        meta_data = [[
            Paragraph(f"<b>Compressor:</b>  {request.compressor_name}", small_s),
            Paragraph(f"<b>Facility:</b>  {request.company_name}", small_s),
            Paragraph(f"<b>Analyst:</b>  {current_user.get('full_name', current_user.get('email','—'))}", small_s),
            Paragraph(f"<b>Date:</b>  {ts}", small_s),
        ]]
        meta_t = Table(meta_data, colWidths=[4.5*cm, 4.5*cm, 5*cm, 4.5*cm])
        meta_t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), colors.HexColor("#f8fafc")),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("LEFTPADDING",  (0,0), (-1,-1), 10),
            ("TOPPADDING",   (0,0), (-1,-1), 7),
            ("BOTTOMPADDING",(0,0), (-1,-1), 7),
            ("LINEBELOW",    (0,0), (-1,-1), 1, C_BORDER),
            ("LINEAFTER",    (0,0), (2,-1),  0.5, C_BORDER),
        ]))
        story.append(meta_t)
        story.append(Spacer(1, 0.5*cm))

        def section_header(title):
            t = Table([[Paragraph(title, h2_s)]], colWidths=[W - 3.6*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), C_BLUE),
                ("LEFTPADDING",  (0,0), (-1,-1), 10),
                ("TOPPADDING",   (0,0), (-1,-1), 6),
                ("BOTTOMPADDING",(0,0), (-1,-1), 6),
                ("LINEBELOW",    (0,0), (-1,-1), 2, C_CYAN),
            ]))
            return t

        def make_table(data, col_widths, green_cells=None, cyan_cells=None):
            t = Table(data, colWidths=col_widths, repeatRows=1)
            style = [
                ("BACKGROUND",   (0,0), (-1,0), C_NAVY),
                ("TEXTCOLOR",    (0,0), (-1,0), C_WHITE),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,0), 9),
                ("ALIGN",        (0,0), (-1,0), "CENTER"),
                ("TOPPADDING",   (0,0), (-1,0), 7),
                ("BOTTOMPADDING",(0,0), (-1,0), 7),
                ("FONTSIZE",     (0,1), (-1,-1), 8.5),
                ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
                ("TEXTCOLOR",    (0,1), (-1,-1), C_TEXT),
                ("TOPPADDING",   (0,1), (-1,-1), 5),
                ("BOTTOMPADDING",(0,1), (-1,-1), 5),
                ("LEFTPADDING",  (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_WHITE, C_LGRAY]),
                ("GRID",         (0,0), (-1,-1), 0.5, C_BORDER),
                ("LINEBELOW",    (0,0), (-1,0),  1.5, C_BLUE),
            ]
            if green_cells:
                for (r_,c_) in green_cells:
                    style += [("TEXTCOLOR",(c_,r_),(c_,r_), C_GREEN),
                              ("FONTNAME", (c_,r_),(c_,r_), "Helvetica-Bold")]
            if cyan_cells:
                for (r_,c_) in cyan_cells:
                    style += [("TEXTCOLOR",(c_,r_),(c_,r_), C_CYAN),
                              ("FONTNAME", (c_,r_),(c_,r_), "Helvetica-Bold")]
            t.setStyle(TableStyle(style))
            return t

        # EXECUTIVE SUMMARY
        story.append(section_header("1.  Executive Summary"))
        story.append(Spacer(1, 0.25*cm))
        story.append(Paragraph(
            f"This report presents the results of an AI-driven optimization analysis performed on compressor "
            f"<b>{request.compressor_name}</b> at <b>{request.company_name}</b>. "
            f"The CompressorAI pipeline — combining DBSCAN clustering, Gradient Boosting Regression (GBR), "
            f"and a Genetic Algorithm optimizer — processed <b>{cs.get('total_points',0)} data points</b> "
            f"({cs.get('clean_points',0)} clean after outlier removal).",
            body_s))
        story.append(Paragraph(
            f"Optimization identified a potential reduction in electrical power consumption from "
            f"<b>{base_elec:.2f} kW</b> (baseline average) to <b>{best_elec:.2f} kW</b> (optimized target), "
            f"representing an <b><font color='#15803d'>{saving:.2f}% energy saving</font></b>. "
            f"Maximum mechanical output achieved: <b>{best_mech:.2f} kW</b>. "
            f"Best Specific Power Consumption (SPC): <b>{best_spc:.4f} kW/(m³/min)</b>.",
            body_s))
        story.append(Spacer(1, 0.35*cm))

        def kpi_card(val, lbl, unit, bg, val_color):
            inner = Table([
                [Paragraph(val,  ParagraphStyle("kv", fontSize=20, textColor=val_color,
                           fontName="Helvetica-Bold", alignment=TA_CENTER, leading=24))],
                [Paragraph(lbl,  ParagraphStyle("kl", fontSize=8,  textColor=C_SUBTEXT,
                           fontName="Helvetica-Bold", alignment=TA_CENTER, leading=11))],
                [Paragraph(unit, ParagraphStyle("ku", fontSize=7,  textColor=C_GRAY,
                           fontName="Helvetica", alignment=TA_CENTER))],
            ], colWidths=[4.2*cm])
            inner.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), bg),
                ("TOPPADDING",   (0,0), (-1,0),  10),
                ("BOTTOMPADDING",(0,2), (-1,2),  10),
                ("TOPPADDING",   (0,1), (-1,2),  2),
                ("BOTTOMPADDING",(0,0), (-1,1),  2),
                ("BOX",          (0,0), (-1,-1), 1, C_BORDER),
                ("LINEABOVE",    (0,0), (-1,0),  3, val_color),
            ]))
            return inner

        kpi_row = [[
            kpi_card(f"{saving:.1f}%",    "ENERGY SAVING",       "vs. baseline",    colors.HexColor("#f0fdf4"), C_GREEN),
            kpi_card(f"{best_elec:.2f}",  "OPTIMAL ELEC. POWER", "kW",              colors.HexColor("#eff6ff"), C_CYAN),
            kpi_card(f"{best_mech:.2f}",  "BEST MECH. POWER",    "kW",              colors.HexColor("#eff6ff"), colors.HexColor("#1d4ed8")),
            kpi_card(f"{best_spc:.4f}",   "BEST SPC",            "kW/(m³/min)",     colors.HexColor("#fefce8"), colors.HexColor("#b45309")),
        ]]
        kpi_t = Table(kpi_row, colWidths=[4.4*cm]*4, hAlign="LEFT")
        kpi_t.setStyle(TableStyle([
            ("LEFTPADDING",  (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING",   (0,0), (-1,-1), 0),
            ("BOTTOMPADDING",(0,0), (-1,-1), 0),
        ]))
        story.append(kpi_t)
        story.append(Spacer(1, 0.5*cm))

        # MODEL PERFORMANCE
        story.append(section_header("2.  Model Performance Scores"))
        story.append(Spacer(1, 0.25*cm))
        # FIX: score keys — engine returns 'convergence', also handle 'ga_convergence'
        convergence_val = scores.get("ga_convergence") or scores.get("convergence", 0)
        score_data = [
            ["Metric", "Score", "Benchmark", "Description"],
            ["DBSCAN Silhouette", f"{scores.get('silhouette',0):.2f}%",  "> 40% Good",  "Cluster quality — density & separation of operating regimes"],
            ["R² Score (GBR)",    f"{scores.get('r2',0):.2f}%",          "> 95% Good",  "Gradient Boosting model accuracy on held-out test data"],
            ["F1 Score",          f"{scores.get('f1',0):.2f}%",          "> 60% Good",  "Classification accuracy for efficiency regime labelling"],
            ["GA Convergence",    f"{convergence_val:.2f}%",             "> 85% Good",  "Genetic Algorithm optimization convergence quality"],
        ]
        story.append(make_table(score_data,
            col_widths=[4.5*cm, 2.5*cm, 2.8*cm, 8.7*cm],
            cyan_cells=[(i,1) for i in range(1,5)]))
        story.append(Spacer(1, 0.5*cm))

        # OPTIMIZATION RESULTS
        story.append(section_header("3.  Optimization Results"))
        story.append(Spacer(1, 0.25*cm))
        res_data = [
            ["Parameter", "Value", "Unit", "Notes"],
            ["Optimal Electrical Power",  f"{best_elec:.3f}", "kW",          "Target operating power after optimization"],
            ["Baseline Electrical Power", f"{base_elec:.3f}", "kW",          "Average power in historical dataset"],
            ["Power Saving",              f"{saving:.2f}%",   "—",           "Reduction from baseline to optimal"],
            ["Best Mechanical Power",     f"{best_mech:.3f}", "kW",          "Maximum shaft output at optimal point"],
            ["Best SPC",                  f"{best_spc:.4f}",  "kW/(m³/min)","Specific Power Consumption at optimum"],
            ["Total Data Points",         str(cs.get("total_points",0)), "rows", "Raw dataset size"],
            ["Clean Data Points",         str(cs.get("clean_points",0)),  "rows", "After DBSCAN outlier removal"],
        ]
        story.append(make_table(res_data,
            col_widths=[5.5*cm, 3*cm, 3*cm, 7*cm],
            green_cells=[(3,1)]))
        story.append(Spacer(1, 0.5*cm))

        if opt_p:
            story.append(section_header("4.  Optimal Operating Parameters"))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(
                "The following parameter ranges define the optimal operating envelope identified by the "
                "Genetic Algorithm. Operating within these bounds is expected to yield maximum energy "
                "efficiency while maintaining required mechanical output.", body_s))
            story.append(Spacer(1, 0.2*cm))
            p_data = [["Parameter", "Optimal Value", "Min Range", "Max Range", "Data Mean", "Unit"]]
            for param, vals in opt_p.items():
                if isinstance(vals, dict):
                    p_data.append([
                        param,
                        f"{vals.get('optimal',0):.4f}",
                        f"{vals.get('min',0):.4f}",
                        f"{vals.get('max',0):.4f}",
                        f"{vals.get('data_mean',0):.4f}" if vals.get('data_mean') else "—",
                        vals.get("unit",""),
                    ])
            story.append(make_table(p_data,
                col_widths=[5.5*cm, 2.8*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.7*cm],
                cyan_cells=[(i,1) for i in range(1,len(p_data))]))
            story.append(Spacer(1, 0.5*cm))

        if fi:
            story.append(section_header("5.  Feature Importance Analysis"))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(
                "Feature importance scores indicate the relative contribution of each operating parameter "
                "to the Gradient Boosting model's predictions.", body_s))
            story.append(Spacer(1, 0.2*cm))
            fi_sorted = sorted(fi.items(), key=lambda x: -x[1])
            fi_data   = [["Rank", "Parameter", "Importance Score", "Importance (%)", "Impact Level"]]
            for idx, (param, score) in enumerate(fi_sorted, 1):
                level = "● High" if score > 0.25 else "◑ Medium" if score > 0.1 else "○ Low"
                fi_data.append([str(idx), param, f"{score:.6f}", f"{score*100:.2f}%", level])
            story.append(make_table(fi_data,
                col_widths=[1.5*cm, 6*cm, 3.2*cm, 3*cm, 4.8*cm],
                cyan_cells=[(1,3)]))
            story.append(Spacer(1, 0.5*cm))

        story.append(section_header("6.  Analysis Configuration Parameters"))
        story.append(Spacer(1, 0.2*cm))
        up_data = [
            ["Parameter", "Value", "Unit", "Description"],
            ["Supply Voltage",         str(up.get("voltage", 415)),         "V",        "3-Phase line voltage"],
            ["Power Factor (cos φ)",   str(up.get("power_factor", 0.9)),    "—",        "Electrical power factor"],
            ["Compression Stages (z)", str(up.get("compression_stages", 2)),"—",        "Number of compression stages"],
            ["P Low",                  str(up.get("p_low", 7.0)),           "bar",      "Lower interpolation pressure"],
            ["P High",                 str(up.get("p_high", 10.0)),         "bar",      "Upper interpolation pressure"],
            ["Q Low",                  str(up.get("q_low", 45.23)),         "m³/min",   "Flow rate at P Low"],
            ["Q High",                 str(up.get("q_high", 35.47)),        "m³/min",   "Flow rate at P High"],
        ]
        story.append(make_table(up_data, col_widths=[4.5*cm, 2.5*cm, 2.5*cm, 9*cm]))
        story.append(Spacer(1, 0.6*cm))

        story.append(Spacer(1, 0.4*cm))
        footer_bg = Table([[""]], colWidths=[W - 3.6*cm])
        footer_bg.setStyle(TableStyle([("LINEABOVE",(0,0),(-1,-1),1,C_BORDER)]))
        story.append(footer_bg)
        footer_data = [[
            Paragraph("⚡ <b>CompressorAI v5.0</b>", ParagraphStyle("fl",fontSize=7.5,textColor=C_NAVY,fontName="Helvetica")),
            Paragraph("DBSCAN Clustering  ·  Gradient Boosting Regression  ·  Genetic Algorithm Optimization",
                      ParagraphStyle("fc",fontSize=7,textColor=C_GRAY,fontName="Helvetica",alignment=TA_CENTER)),
            Paragraph(f"Confidential  ·  {ts_f}",
                      ParagraphStyle("fr",fontSize=7.5,textColor=C_GRAY,fontName="Helvetica",alignment=TA_RIGHT)),
        ]]
        footer_t = Table(footer_data, colWidths=[4.5*cm, 9.5*cm, 4.5*cm])
        footer_t.setStyle(TableStyle([
            ("TOPPADDING",  (0,0),(-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        story.append(footer_t)

        doc = SimpleDocTemplate(buf, pagesize=A4,
            topMargin=1.5*cm, bottomMargin=1.5*cm,
            leftMargin=1.8*cm, rightMargin=1.8*cm)
        doc.build(story)
        buf.seek(0)
        import re
        _safe = re.sub(r'[^\x00-\x7F]', '', request.compressor_name).replace(' ','_').strip('_') or 'report'
        fname = f"CompressorAI_{_safe}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed. Run: pip install reportlab")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


# ── Excel Report ──────────────────────────────────────────────
@router.post("/excel")
async def generate_excel_report(request: ReportRequest, current_user=Depends(get_current_user)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        r  = request.analysis_results
        up = request.user_params or {}
        ts = datetime.now().strftime("%d %B %Y  %H:%M")

        NAVY     = "0a1628"
        BLUE     = "1a3a6b"
        WHITE    = "ffffff"
        LGRAY    = "f1f5f9"
        GREEN    = "15803d"
        GREEN_BG = "f0fdf4"
        CYAN     = "0066cc"
        GRAY     = "64748b"
        TEXT     = "1e293b"
        BORDER_C = "cbd5e1"

        def fill(hex_): return PatternFill("solid", fgColor=hex_)
        def font(hex_, bold=False, sz=10): return Font(color=hex_, bold=bold, size=sz, name="Calibri")
        thin    = Side(style="thin", color=BORDER_C)
        bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
        bdr_hdr = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color=CYAN))

        def write_header_row(ws, row, cols):
            for ci, txt in enumerate(cols, 1):
                c = ws.cell(row=row, column=ci, value=txt)
                c.font      = font(WHITE, bold=True, sz=10)
                c.fill      = fill(NAVY)
                c.border    = bdr_hdr
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def write_data_row(ws, row, vals, alt=False, green_cols=None, cyan_cols=None):
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill   = fill(LGRAY if alt else WHITE)
                c.border = bdr
                c.alignment = Alignment(horizontal="left" if ci==1 else "center",
                                        vertical="center", wrap_text=False)
                if green_cols and ci in green_cols:
                    c.font = font(GREEN, bold=True, sz=10)
                elif cyan_cols and ci in cyan_cols:
                    c.font = font(CYAN, bold=True, sz=10)
                else:
                    c.font = font(TEXT, sz=10)

        def set_col_widths(ws, widths):
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        def section_title(ws, row, title, ncols=4):
            c = ws.cell(row=row, column=1, value=title)
            c.font      = font(WHITE, bold=True, sz=11)
            c.fill      = fill(BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = Border(bottom=Side(style="medium", color=CYAN))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            ws.row_dimensions[row].height = 22

        saving    = r.get("power_saving_percent", 0) or 0
        best_elec = r.get("best_electrical_power", 0) or 0
        best_mech = r.get("best_mechanical_power", 0) or 0
        base_elec = r.get("baseline_electrical_power", 0) or 0
        best_spc  = r.get("best_spc", 0) or 0
        scores    = r.get("scores", {})
        cs        = r.get("cluster_stats", {})

        ws1 = wb.active
        ws1.title = "Summary"
        ws1.sheet_properties.tabColor = "facc15"
        ws1.freeze_panes = "A4"

        ws1.merge_cells("A1:F1")
        c = ws1["A1"]
        c.value     = "CompressorAI v5  —  Industrial Air Compressor Optimization Report"
        c.font      = font("facc15", bold=True, sz=15)
        c.fill      = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.row_dimensions[1].height = 32

        ws1.merge_cells("A2:F2")
        c2 = ws1["A2"]
        c2.value     = f"Compressor: {request.compressor_name}   |   Facility: {request.company_name}   |   Generated: {ts}"
        c2.font      = font(GRAY, sz=9)
        c2.fill      = fill("f8fafc")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.row_dimensions[2].height = 18

        set_col_widths(ws1, [30, 18, 18, 18, 18, 18])

        row = 4
        section_title(ws1, row, "KEY PERFORMANCE INDICATORS", 6)
        row += 1
        kpis = [
            ("Energy Saving",           f"{saving:.2f}%",             True,  False),
            ("Optimal Electrical Power", f"{best_elec:.3f} kW",       False, True),
            ("Best Mechanical Power",    f"{best_mech:.3f} kW",       False, False),
            ("Baseline Electrical",      f"{base_elec:.3f} kW",       False, False),
            ("Best SPC",                 f"{best_spc:.4f} kW/m³/min", False, True),
            ("Clean Data Points",        f"{cs.get('clean_points',0)} / {cs.get('total_points',0)}", False, False),
        ]
        write_header_row(ws1, row, ["Metric", "Value", "Metric", "Value", "Metric", "Value"])
        row += 1
        for i in range(0, len(kpis), 2):
            left  = kpis[i]   if i   < len(kpis) else ("", "", False, False)
            right = kpis[i+1] if i+1 < len(kpis) else ("", "", False, False)
            alt   = (row % 2 == 0)
            for ci, (lbl, val, grn, cyn) in enumerate([(left[0],left[1],left[2],left[3]),
                                                        (right[0],right[1],right[2],right[3])], 0):
                c1 = ws1.cell(row=row, column=ci*3+1, value=lbl)
                c1.font = font(GRAY, sz=9); c1.fill = fill(LGRAY if alt else WHITE); c1.border = bdr
                c2_ = ws1.cell(row=row, column=ci*3+2, value=val)
                c2_.fill = fill(GREEN_BG if grn else (LGRAY if alt else WHITE)); c2_.border = bdr
                c2_.alignment = Alignment(horizontal="center", vertical="center")
                if grn:   c2_.font = font(GREEN, bold=True, sz=11)
                elif cyn: c2_.font = font(CYAN,  bold=True, sz=11)
                else:     c2_.font = font(TEXT,  bold=True, sz=10)
                ws1.cell(row=row, column=ci*3+3).fill = fill(LGRAY if alt else WHITE)
            row += 1

        row += 1
        section_title(ws1, row, "MODEL PERFORMANCE SCORES", 6)
        row += 1
        write_header_row(ws1, row, ["Model Component", "Score", "Benchmark", "Description", "", ""])
        ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        row += 1
        convergence_val = scores.get("ga_convergence") or scores.get("convergence", 0)
        score_rows = [
            ("DBSCAN Silhouette", f"{scores.get('silhouette',0):.2f}%",  "> 40% Good",  "Cluster quality — density & separation"),
            ("R² Score (GBR)",    f"{scores.get('r2',0):.2f}%",          "> 95% Good",  "GBR model accuracy on test data"),
            ("F1 Score",          f"{scores.get('f1',0):.2f}%",          "> 60% Good",  "Efficiency classification accuracy"),
            ("GA Convergence",    f"{convergence_val:.2f}%",             "> 85% Good",  "Optimization convergence quality"),
        ]
        for i, (m, s, b, d) in enumerate(score_rows):
            alt = (i % 2 == 0)
            write_data_row(ws1, row, [m, s, b, d, "", ""], alt=alt, cyan_cols=[2])
            ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            row += 1

        # Sheet 2 — Optimal Parameters
        ws2 = wb.create_sheet("Optimal Parameters")
        ws2.sheet_properties.tabColor = "0066cc"
        ws2.freeze_panes = "A3"
        banner2 = ws2["A1"]
        banner2.value     = f"Optimal Operating Parameters  —  {request.compressor_name}"
        banner2.font      = font(WHITE, bold=True, sz=12)
        banner2.fill      = fill(NAVY)
        banner2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.merge_cells("A1:F1")
        ws2.row_dimensions[1].height = 26
        set_col_widths(ws2, [28, 14, 14, 14, 14, 10])
        write_header_row(ws2, 2, ["Parameter", "Optimal Value", "Min Range", "Max Range", "Data Mean", "Unit"])
        row2 = 3
        for param, vals in r.get("optimal_parameters", {}).items():
            if isinstance(vals, dict):
                write_data_row(ws2, row2,
                    [param,
                     round(vals.get("optimal",0), 4),
                     round(vals.get("min",0), 4),
                     round(vals.get("max",0), 4),
                     round(vals.get("data_mean",0), 4) if vals.get("data_mean") else "—",
                     vals.get("unit","")],
                    alt=(row2%2==0), cyan_cols=[2])
                row2 += 1

        # Sheet 3 — Feature Importance
        ws3 = wb.create_sheet("Feature Importance")
        ws3.sheet_properties.tabColor = "15803d"
        ws3.freeze_panes = "A3"
        banner3 = ws3["A1"]
        banner3.value     = f"Feature Importance Analysis  —  {request.compressor_name}"
        banner3.font      = font(WHITE, bold=True, sz=12)
        banner3.fill      = fill(NAVY)
        banner3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.merge_cells("A1:E1")
        ws3.row_dimensions[1].height = 26
        set_col_widths(ws3, [28, 8, 16, 16, 14])
        write_header_row(ws3, 2, ["Parameter", "Rank", "Importance Score", "Importance (%)", "Impact Level"])
        fi_sorted = sorted(r.get("feature_importance", {}).items(), key=lambda x: -x[1])
        for ri, (feat, imp) in enumerate(fi_sorted, 3):
            level = "High" if imp > 0.25 else "Medium" if imp > 0.1 else "Low"
            write_data_row(ws3, ri,
                [feat, ri-2, round(imp,6), f"{imp*100:.2f}%", level],
                alt=(ri%2==0),
                green_cols=[3,4] if imp > 0.25 else None)

        # Sheet 4 — Configuration
        ws4 = wb.create_sheet("Configuration")
        ws4.sheet_properties.tabColor = "64748b"
        banner4 = ws4["A1"]
        banner4.value     = "Analysis Configuration Parameters"
        banner4.font      = font(WHITE, bold=True, sz=12)
        banner4.fill      = fill(NAVY)
        banner4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws4.merge_cells("A1:D1")
        ws4.row_dimensions[1].height = 26
        set_col_widths(ws4, [28, 14, 12, 30])
        write_header_row(ws4, 2, ["Parameter", "Value", "Unit", "Description"])
        cfg_rows = [
            ("Supply Voltage",         up.get("voltage",415),         "V",       "3-Phase line voltage"),
            ("Power Factor (cos φ)",   up.get("power_factor",0.9),    "—",       "Electrical power factor"),
            ("Compression Stages (z)", up.get("compression_stages",2),"—",       "Number of compression stages"),
            ("P Low",                  up.get("p_low",7.0),           "bar",     "Lower interpolation pressure"),
            ("P High",                 up.get("p_high",10.0),         "bar",     "Upper interpolation pressure"),
            ("Q Low",                  up.get("q_low",45.23),         "m³/min",  "Flow rate at P Low"),
            ("Q High",                 up.get("q_high",35.47),        "m³/min",  "Flow rate at P High"),
        ]
        for ri, row_data in enumerate(cfg_rows, 3):
            write_data_row(ws4, ri, list(row_data), alt=(ri%2==0))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        import re
        _safe = re.sub(r'[^\x00-\x7F]', '', request.compressor_name).replace(' ','_').strip('_') or 'report'
        fname = f"CompressorAI_{_safe}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")


# ── My Reports ─────────────────────────────────────────────────
@router.get("/my")
async def get_my_reports(current_user=Depends(get_current_user)):
    supabase = get_supabase_client()
    # FIX: analyses → analysis_results, compressor_id → unit_id
    res = supabase.table("analysis_results").select(
        "id,unit_id,dataset_id,power_saving_percent,"
        "best_electrical_power,scores,created_at"
    ).eq("user_id", current_user["sub"]).order("created_at", desc=True).execute()
    return res.data or []